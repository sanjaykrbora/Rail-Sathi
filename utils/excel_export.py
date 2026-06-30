from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment


BASE_DIR = Path(__file__).resolve().parent.parent
REPORT_DIR = BASE_DIR / "reports"

REPORT_DIR.mkdir(exist_ok=True)


def export_excel(title, dataframe, filename):

    workbook = Workbook()

    sheet = workbook.active

    sheet.title = title

    title_cell = sheet["A1"]
    title_cell.value = "Rail Sathi"
    title_cell.font = Font(
        size=18,
        bold=True
    )

    sheet["A2"] = "N.F. Railway Mechanical Workshop, Dibrugarh"

    row = 4

    for column, name in enumerate(dataframe.columns, start=1):

        cell = sheet.cell(
            row=row,
            column=column
        )

        cell.value = name

        cell.font = Font(
            bold=True,
            color="FFFFFF"
        )

        cell.fill = PatternFill(
            fill_type="solid",
            fgColor="0B5394"
        )

        cell.alignment = Alignment(
            horizontal="center"
        )

    for r, values in enumerate(
        dataframe.values.tolist(),
        start=5
    ):

        for c, value in enumerate(values, start=1):

            sheet.cell(
                row=r,
                column=c
            ).value = value

    for column_cells in sheet.columns:

        length = max(
            len(str(cell.value))
            if cell.value else 0
            for cell in column_cells
        )

        sheet.column_dimensions[
            column_cells[0].column_letter
        ].width = length + 5

    file_path = REPORT_DIR / filename

    workbook.save(file_path)

    return file_path